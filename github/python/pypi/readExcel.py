# pypi => Python Package Index
# for python3 use pip3 to install any packages, Ex: sudo pip3 install openpyxl
from openpyxl import load_workbook as loadWorkBook

workBook = loadWorkBook('transactions.xlsx')
firstSheet = workBook['Sheet1']
getFirstCell = firstSheet['a1'] # we can even write it lie firstSheet.cell(1,1)

# get max row for that first sheet
maxRow = firstSheet.max_row
# we don't need first cell so start loop from 2nd index
for row in range(2, maxRow + 1): 
    cell = firstSheet.cell(row, 3).value
    #  after updating price by 10 insert them into sheet
    updatedVal = 10 * cell
    correctedVal = firstSheet.cell(row, 4)
    correctedVal.value = updatedVal

# saving changes to new file
workBook.save('newFile.xlsx')